import csv
import json
from decimal import Decimal
from .models import Entrada, Cliente, Venta, DetalleVenta, MovimientoInventario, MovimientoProducto, PreInventario, Abono

from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Q
from django.core.paginator import Paginator

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


from .models import Pedido, DetallePedido
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Entrada, Cliente, Venta, DetalleVenta, MovimientoInventario

from core.models import PreInventario

from django.utils import timezone

#from django.db.models import Sum, F, Q
from django.db.models.functions import Coalesce

from .models import Entrada, Cliente, Venta, DetalleVenta, MovimientoInventario, MovimientoProducto

from core.models import Cliente


from core.models import Proveedor

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak

from django.db import models
from weasyprint import HTML
from django.template.loader import get_template

import json
from django.core.serializers.json import DjangoJSONEncoder
from core.decorators import grupo_requerido, administrador_requerido
from django.contrib.auth.models import User

def login_view(request):
    # Crear superusuario si no existe (solo para producción en Render)
    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        print("Superusuario admin creado automáticamente")
    
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Credenciales incorrectas')
    return render(request, 'core/login.html')

# ==================== AUTENTICACIÓN ====================

def login_view(request):
    # Crear superusuario si no existe (solo para producción en Render)
    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123'
        )
        print("Superusuario admin creado automáticamente")
    
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Credenciales incorrectas')
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

# ==================== HOME ====================
@login_required
def home(request):
    try:
        rol = request.user.empleado.rol
    except:
        rol = 'admin'
    
    # Contar pedidos pendientes de TODOS los usuarios
    pendientes_count = Pedido.objects.filter(estado='pendiente').count()
    
    context = {
        'rol': rol,
        'nombre_usuario': request.user.get_full_name() or request.user.username,
        'pendientes_count': pendientes_count,
    }
    return render(request, 'core/home.html', context)
# ==================== ENTRADAS ====================

@login_required
@administrador_requerido  # Solo administradores
def listar_entradas(request):
    entradas_list = Entrada.objects.all().order_by('-fecha_registro')
    
    # Filtros
    busqueda = request.GET.get('busqueda', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if busqueda:
        entradas_list = entradas_list.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda) |
            Q(no_duca__icontains=busqueda) |
            Q(ubicacion__icontains=busqueda)
        )
    
    if fecha_desde:
        entradas_list = entradas_list.filter(fecha_ingreso__gte=fecha_desde)
    if fecha_hasta:
        entradas_list = entradas_list.filter(fecha_ingreso__lte=fecha_hasta)
    
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [20, 50, 100]:
            por_pagina = 20
    except:
        por_pagina = 20
    
    paginator = Paginator(entradas_list, por_pagina)
    page_number = request.GET.get('page')
    entradas = paginator.get_page(page_number)
    
    # Importar modelo Categoria
    from core.models import Categoria
    
    context = {
        'entradas': entradas,
        'busqueda': busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
        'proveedores': Proveedor.objects.all().order_by('nombre'),
        'categorias': Categoria.objects.all().order_by('nombre'),  # Agrega esta línea
    }
    return render(request, 'core/entradas.html', context)



@login_required
def crear_entrada(request):
    if request.method == 'POST':
        try:
            # Convertir cantidad a entero ANTES de crear la entrada
              # Validar campos obligatorios
            codigo = request.POST.get('codigo', '').strip()
            producto = request.POST.get('producto', '').strip()
            categoria = request.POST.get('categoria', '').strip()
            cantidad = request.POST.get('cantidad')
            precio_compra = request.POST.get('precio_compra')
            fecha_ingreso = request.POST.get('fecha_ingreso')
            ubicacion = request.POST.get('ubicacion', '').strip()
            cantidad = int(request.POST.get('cantidad'))
            precio_compra = float(request.POST.get('precio_compra'))
            
            proveedor_texto = request.POST.get('proveedor')

             
            if not codigo:
                return JsonResponse({'success': False, 'error': 'El código es obligatorio'})
            if not producto:
                return JsonResponse({'success': False, 'error': 'El producto es obligatorio'})
            if not categoria:
                return JsonResponse({'success': False, 'error': 'La categoría es obligatoria'})
            if not cantidad:
                return JsonResponse({'success': False, 'error': 'La cantidad es obligatoria'})
            if not fecha_ingreso:
                return JsonResponse({'success': False, 'error': 'La fecha de ingreso es obligatoria'})
            if not ubicacion:
                return JsonResponse({'success': False, 'error': 'La ubicación es obligatoria'})
            
            cantidad = int(cantidad)
            precio_compra = float(precio_compra)
            
            # ========== SINCRONIZAR CON MÓDULO PROVEEDORES ==========
            if proveedor_texto:
                from core.models import Proveedor
                proveedor_existente = Proveedor.objects.filter(nombre=proveedor_texto).first()
                if not proveedor_existente:
                    Proveedor.objects.create(
                        nit=f"TEMP-{proveedor_texto[:10]}",
                        nombre=proveedor_texto,
                        pais="Guatemala",
                        observaciones="Creado automáticamente desde Entradas"
                    )
            # ======================================================
            
            entrada = Entrada.objects.create(
                producto=request.POST.get('producto').lower(),
                codigo=request.POST.get('codigo').lower(),
                no_duca=request.POST.get('no_duca'),
                categoria=request.POST.get('categoria').lower(),
                cantidad=cantidad,
                precio_compra=precio_compra,
                proveedor=proveedor_texto,
                ubicacion=request.POST.get('ubicacion').lower() if request.POST.get('ubicacion') else '',
                fecha_ingreso=request.POST.get('fecha_ingreso'),
                comentario=request.POST.get('comentario'),
                imagen=request.FILES.get('imagen'),
                usuario=request.user
            )
            
            # Actualizar PreInventario
            pre = PreInventario.objects.filter(
                codigo=entrada.codigo,
                producto=entrada.producto,
                categoria=entrada.categoria,
                proveedor=entrada.proveedor or '',
                ubicacion=entrada.ubicacion
            ).first()
            
            if pre:
                pre.cantidad += cantidad
                pre.save()
            else:
                PreInventario.objects.create(
                    codigo=entrada.codigo,
                    producto=entrada.producto,
                    categoria=entrada.categoria,
                    proveedor=entrada.proveedor or '',
                    cantidad=cantidad,
                    ubicacion=entrada.ubicacion or 'Sin ubicación'
                )
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@login_required
def editar_entrada(request, id):
    entrada = get_object_or_404(Entrada, id=id)
    
    if request.method == 'GET':
        data = {
            'id': entrada.id,
            'producto': entrada.producto,
            'codigo': entrada.codigo,
            'no_duca': entrada.no_duca or '',
            'categoria': entrada.categoria,
            'cantidad': entrada.cantidad,
            'precio_compra': str(entrada.precio_compra),
            'proveedor': entrada.proveedor or '',
            'ubicacion': entrada.ubicacion or '',
            'fecha_ingreso': entrada.fecha_ingreso.strftime('%Y-%m-%d'),
            'comentario': entrada.comentario or '',
            'imagen_url': entrada.imagen.url if entrada.imagen else '',
        }
        return JsonResponse(data)
    
    if request.method == 'POST':
        try:
            proveedor_texto = request.POST.get('proveedor', '').strip()
            
            # Convertir cantidad a entero
            cantidad = int(request.POST.get('cantidad'))
            
            # Sincronizar con módulo Proveedores
            if proveedor_texto:
                proveedor_existente = Proveedor.objects.filter(nombre=proveedor_texto).first()
                if not proveedor_existente:
                    Proveedor.objects.create(
                        nit=f"TEMP-{proveedor_texto[:10]}",
                        nombre=proveedor_texto,
                        pais="Guatemala",
                        observaciones="Creado automáticamente desde Entradas"
                    )
                    print(f"Proveedor creado automáticamente: {proveedor_texto}")
            
            entrada.producto = request.POST.get('producto').lower()
            entrada.codigo = request.POST.get('codigo').lower()
            entrada.no_duca = request.POST.get('no_duca')
            entrada.categoria = request.POST.get('categoria').lower()
            entrada.cantidad = cantidad  # Usar la variable convertida
            entrada.precio_compra = request.POST.get('precio_compra')
            entrada.proveedor = proveedor_texto
            entrada.ubicacion = request.POST.get('ubicacion').lower() if request.POST.get('ubicacion') else ''
            entrada.fecha_ingreso = request.POST.get('fecha_ingreso')
            entrada.comentario = request.POST.get('comentario')
            if request.FILES.get('imagen'):
                entrada.imagen = request.FILES.get('imagen')
            entrada.save()
            
            # Actualizar PreInventario
            pre = PreInventario.objects.filter(
                codigo=entrada.codigo,
                producto=entrada.producto,
                categoria=entrada.categoria,
                proveedor=entrada.proveedor or '',
                ubicacion=entrada.ubicacion
            ).first()
            
            if pre:
                pre.cantidad = cantidad  # Usar la variable convertida
                pre.save()
            else:
                PreInventario.objects.create(
                    codigo=entrada.codigo,
                    producto=entrada.producto,
                    categoria=entrada.categoria,
                    proveedor=entrada.proveedor or '',
                    cantidad=cantidad,
                    ubicacion=entrada.ubicacion or 'Sin ubicación'
                )
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def eliminar_entrada(request, id):
    if request.method == 'POST':
        entrada = get_object_or_404(Entrada, id=id)
        entrada.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


#-----------------------EXPORTAR PDF Y EXCEL---------
@login_required
def exportar_reporte(request, formato):
    entradas_list = Entrada.objects.all().order_by('-fecha_registro')
    
    busqueda = request.GET.get('busqueda', '')
    fecha_ingreso = request.GET.get('fecha_ingreso', '')
    
    if busqueda:
        entradas_list = entradas_list.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    if fecha_ingreso:
        entradas_list = entradas_list.filter(fecha_ingreso=fecha_ingreso)
    
    if formato == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entradas"
        
        headers = ['Código', 'Producto', 'Categoría', 'Cantidad', 'Precio Compra', 'Proveedor', 'Fecha Ingreso', 'Descripción', 'Comentario']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, entrada in enumerate(entradas_list, 2):
            ws.cell(row=row, column=1, value=entrada.codigo)
            ws.cell(row=row, column=2, value=entrada.producto)
            ws.cell(row=row, column=3, value=entrada.categoria)
            ws.cell(row=row, column=4, value=entrada.cantidad)
            ws.cell(row=row, column=5, value=float(entrada.precio_compra))
            ws.cell(row=row, column=6, value=entrada.proveedor or '')
            ws.cell(row=row, column=7, value=entrada.fecha_ingreso.strftime('%d/%m/%Y'))
            #ws.cell(row=row, column=8, value=entrada.descripcion or '')
            ws.cell(row=row, column=8, value=entrada.comentario or '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="entradas.xlsx"'
        wb.save(response)
        return response
    
    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="entradas.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Entradas", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(" ", styles['Normal']))
        
        data = [['Código', 'Producto', 'Categoría', 'Cantidad', 'Precio', 'Proveedor', 'Fecha']]
        
        for entrada in entradas_list:
            data.append([
                entrada.codigo, entrada.producto, entrada.categoria, str(entrada.cantidad),
                f"Q{entrada.precio_compra}", entrada.proveedor or '', entrada.fecha_ingreso.strftime('%d/%m/%Y')
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
    
    return JsonResponse({'error': 'Formato no soportado'}, status=400)



@login_required
def exportar_productos_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    
    # Obtener productos con ubicación (sin agrupar)
    productos = obtener_productos_filtrados(request)
    
    resultados = []
    for p in productos:
        if p.cantidad > 0:
            stock = p.cantidad
            if stock <= 2:
                estado = "Crítico"
            elif stock <= 6:
                estado = "Bajo"
            else:
                estado = "Normal"
            
            resultados.append({
                'codigo': p.codigo,
                'producto': p.producto,
                'categoria': p.categoria,
                'cantidad': stock,
                'proveedor': p.proveedor or '-',
                'ubicacion': p.ubicacion,
                'estado': estado
            })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Reporte de Productos - Inventario Actual", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(" ", styles['Normal']))
    
    # Mostrar filtros aplicados
    busqueda = request.GET.get('busqueda', '')
    ubicacion = request.GET.get('ubicacion', '')
    filtros = []
    if busqueda:
        filtros.append(f"Búsqueda: {busqueda}")
    if ubicacion:
        filtros.append(f"Ubicación: {ubicacion}")
    if filtros:
        elements.append(Paragraph(f"<b>Filtros:</b> {', '.join(filtros)}", styles['Normal']))
        elements.append(Paragraph(" ", styles['Normal']))
    
    data = [['Código', 'Producto', 'Categoría', 'Cantidad', 'Estado', 'Proveedor', 'Ubicación']]
    for r in resultados:
        data.append([r['codigo'], r['producto'], r['categoria'], str(r['cantidad']), r['estado'], r['proveedor'], r['ubicacion']])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_productos_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Obtener productos con ubicación (sin agrupar)
    productos = obtener_productos_filtrados(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    headers = ['Código', 'Producto', 'Categoría', 'Cantidad', 'Estado', 'Proveedor', 'Ubicación']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, p in enumerate(productos, 2):
        if p.cantidad > 0:
            stock = p.cantidad
            if stock <= 2:
                estado = "Crítico"
            elif stock <= 6:
                estado = "Bajo"
            else:
                estado = "Normal"
            
            ws.cell(row=row, column=1, value=p.codigo)
            ws.cell(row=row, column=2, value=p.producto)
            ws.cell(row=row, column=3, value=p.categoria)
            ws.cell(row=row, column=4, value=stock)
            ws.cell(row=row, column=5, value=estado)
            ws.cell(row=row, column=6, value=p.proveedor or '-')
            ws.cell(row=row, column=7, value=p.ubicacion)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


def obtener_productos_filtrados(request):
    from django.db.models import Q
    
    # Obtener productos con su ubicación (sin agrupar)
    productos = PreInventario.objects.all().order_by('codigo', 'producto', 'ubicacion')
    
    # Filtro de búsqueda
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    # Filtro de ubicación
    ubicacion = request.GET.get('ubicacion', '')
    if ubicacion:
        productos = productos.filter(ubicacion=ubicacion)
    
    return productos
# ==================== VENTAS ====================




@login_required
def buscar_productos_venta(request):
    busqueda = request.GET.get('q', '')
    
    productos = PreInventario.objects.all()
    
    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    resultados = []
    for p in productos:
        if p.cantidad > 0:
            resultados.append({
                'codigo': p.codigo,
                'producto': p.producto,
                'proveedor': p.proveedor or '',
                'precio': 0,  # Precio en cero, se agregará manualmente
                'stock': p.cantidad,
                'ubicacion': p.ubicacion
            })
    
    return JsonResponse(resultados, safe=False)



@login_required
def buscar_cliente(request):
    busqueda = request.GET.get('q', '')
    
    clientes = Cliente.objects.filter(
        Q(nit__icontains=busqueda) | 
        Q(nombre__icontains=busqueda)
    )[:10]
    
    resultados = [{
        'id': c.id,
        'nit': c.nit or '',  # Si es NULL, enviar vacío
        'nombre': c.nombre,
        'direccion': c.direccion or ''
    } for c in clientes]
    
    return JsonResponse(resultados, safe=False)



@login_required
def crear_cliente_rapido(request):
    if request.method == 'POST':
        try:
            nit = request.POST.get('nit')
            # Si el NIT está vacío, guardar como NULL
            if nit == '' or nit == 'CF' or nit == 'C/F':
                nit = None
            
            cliente = Cliente.objects.create(
                nit=nit,
                nombre=request.POST.get('nombre'),
                direccion=request.POST.get('direccion'),
                telefono=request.POST.get('telefono'),
                observaciones=request.POST.get('observaciones')
            )
            return JsonResponse({
                'success': True,
                'id': cliente.id,
                'nit': cliente.nit or '',
                'nombre': cliente.nombre
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def finalizar_venta(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            cliente_id = data.get('cliente_id')
            cliente = Cliente.objects.get(id=cliente_id) if cliente_id else None
            
            ultima_venta = Venta.objects.order_by('-id').first()
            num_factura = data.get('numero_factura') or str(int(ultima_venta.numero_factura) + 1).zfill(8) if ultima_venta else '00000001'
            
            venta = Venta.objects.create(
                numero_factura=num_factura,
                cliente=cliente,
                subtotal=data['subtotal'],
                total=data['total'],
                tipo_documento=data['tipo_documento'],
                tipo_pago=data.get('tipo_pago', 'contado'),
                comentario=data.get('comentario_venta', ''),
                usuario=request.user
            )
            
            for item in data['carrito']:
                DetalleVenta.objects.create(
                    venta=venta,
                    producto=item['producto'],
                    codigo=item['codigo'],
                    proveedor=item.get('proveedor', ''),
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio']
                )
                
                # Buscar el producto en PreInventario por código, producto y ubicación específica
                pre = PreInventario.objects.filter(
                    codigo=item['codigo'],
                    producto=item['producto'],
                    ubicacion=item.get('ubicacion', '')
                ).first()
                
                if not pre:
                    # Si no se encuentra en la ubicación específica, buscar en cualquier ubicación
                    pre = PreInventario.objects.filter(
                        codigo=item['codigo'],
                        producto=item['producto']
                    ).first()
                
                if pre:
                    if pre.cantidad >= item['cantidad']:
                        pre.cantidad -= item['cantidad']
                        if pre.cantidad == 0:
                            pre.delete()
                        else:
                            pre.save()
                    else:
                        return JsonResponse({'success': False, 'error': f'Stock insuficiente para {item["producto"]}. Disponible: {pre.cantidad}'})
                else:
                    return JsonResponse({'success': False, 'error': f'Producto {item["producto"]} no encontrado en inventario'})
            
            return JsonResponse({'success': True, 'factura': num_factura})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
@grupo_requerido('Administrador', 'Operativo')
def ventas(request):
    context = {
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/ventas.html', context)


@login_required
@grupo_requerido('Administrador', 'Operativo')
def historial_ventas(request):
    ventas_list = Venta.objects.all().order_by('-fecha')
    
    busqueda = request.GET.get('busqueda', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if busqueda:
        ventas_list = ventas_list.filter(
            Q(numero_factura__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(cliente__nit__icontains=busqueda)
        )
    
    if fecha_desde:
        ventas_list = ventas_list.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        ventas_list = ventas_list.filter(fecha__date__lte=fecha_hasta)
    
    # Calcular saldo para cada venta
    from decimal import Decimal
    from django.db.models import Sum
    
    for venta in ventas_list:
        if venta.tipo_pago == 'credito':
            abonos_total = Abono.objects.filter(venta=venta).aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
            venta.saldo = float(venta.total) - float(abonos_total)
            if venta.saldo < 0:
                venta.saldo = 0
        else:
            venta.saldo = 0
    
    por_pagina = request.GET.get('por_pagina', 20)
    paginator = Paginator(ventas_list, por_pagina)
    page_number = request.GET.get('page')
    ventas_paginadas = paginator.get_page(page_number)
    
    context = {
        'ventas': ventas_paginadas,
        'busqueda': busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/historial_ventas.html', context)


@login_required
def listar_productos(request):
    productos = PreInventario.objects.filter(cantidad__gt=0).order_by('codigo', 'producto', 'ubicacion')
    
    # Filtros
    busqueda = request.GET.get('busqueda', '')
    ubicacion_filtro = request.GET.get('ubicacion', '')
    
    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    if ubicacion_filtro:
        productos = productos.filter(ubicacion=ubicacion_filtro)
    
    # Obtener ubicaciones únicas
    ubicaciones = PreInventario.objects.values_list('ubicacion', flat=True).distinct()
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    paginator = Paginator(productos, por_pagina)
    page_number = request.GET.get('page')
    productos_paginados = paginator.get_page(page_number)
    
    # Agregar estado
    for p in productos_paginados:
        if p.cantidad <= 2:
            p.estado_color = 'danger'
            p.estado_texto = 'Crítico'
        elif p.cantidad <= 6:
            p.estado_color = 'warning'
            p.estado_texto = 'Bajo'
        else:
            p.estado_color = 'success'
            p.estado_texto = 'Normal'
    
    context = {
        'productos': productos_paginados,
        'busqueda': busqueda,
        'ubicacion_filtro': ubicacion_filtro,
        'ubicaciones': ubicaciones,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/productos.html', context)
#-------------- MOVIEMINTO PRODUCTOS------------#


def calcular_stock_producto(codigo, producto, categoria, proveedor, ubicacion):
    # Stock desde entradas (histórico)
    entradas = Entrada.objects.filter(
        codigo=codigo,
        producto=producto,
        categoria=categoria,
        proveedor=proveedor,
        ubicacion=ubicacion
    ).aggregate(total=Coalesce(Sum('cantidad'), 0))['total']
    
    # Ventas (salen de cualquier ubicación)
    ventas = DetalleVenta.objects.filter(
        codigo=codigo,
        producto=producto
    ).aggregate(total=Coalesce(Sum('cantidad'), 0))['total']
    
    # Movimientos de producto que afectan esta ubicación
    movimientos_origen = MovimientoProducto.objects.filter(
        producto_codigo=codigo,
        producto_nombre=producto,
        categoria=categoria,
        proveedor=proveedor,
        ubicacion_origen=ubicacion,
        tipo='cambio_ubicacion'
    ).aggregate(total=Coalesce(Sum('cantidad'), 0))['total']
    
    movimientos_destino = MovimientoProducto.objects.filter(
        producto_codigo=codigo,
        producto_nombre=producto,
        categoria=categoria,
        proveedor=proveedor,
        ubicacion_destino=ubicacion,
        tipo='cambio_ubicacion'
    ).aggregate(total=Coalesce(Sum('cantidad'), 0))['total']
    
    # Ajustes directos
    ajustes = MovimientoProducto.objects.filter(
        producto_codigo=codigo,
        producto_nombre=producto,
        categoria=categoria,
        proveedor=proveedor,
        ubicacion_destino=ubicacion,
        tipo='ajuste'
    ).order_by('-fecha').first()
    
    stock = entradas - ventas - movimientos_origen + movimientos_destino
    
    # Si hay ajuste reciente, sobrescribe el stock
    if ajustes:
        stock = ajustes.cantidad
    
    return stock

@login_required
def registrar_movimiento_producto(request):
    if request.method == 'POST':
        try:
            codigo = request.POST.get('producto_codigo')
            producto = request.POST.get('producto_nombre')
            categoria = request.POST.get('categoria')
            proveedor = request.POST.get('proveedor')
            cantidad = int(request.POST.get('cantidad'))
            ubicacion_origen = request.POST.get('ubicacion_origen')
            ubicacion_destino = request.POST.get('ubicacion_destino')
            motivo = request.POST.get('motivo')
            
            # Guardar en historial de movimientos
            MovimientoProducto.objects.create(
                producto_codigo=codigo,
                producto_nombre=producto,
                categoria=categoria,
                proveedor=proveedor,
                cantidad=cantidad,
                ubicacion_origen=ubicacion_origen,
                ubicacion_destino=ubicacion_destino,
                tipo='cambio_ubicacion',
                motivo=motivo,
                usuario=request.user
            )
            
            # 1. Restar del origen en PreInventario
            if ubicacion_origen:
                origen = PreInventario.objects.filter(
                    codigo=codigo,
                    producto=producto,
                    categoria=categoria,
                    proveedor=proveedor,
                    ubicacion=ubicacion_origen
                ).first()
                
                if origen:
                    if origen.cantidad >= cantidad:
                        origen.cantidad -= cantidad
                        if origen.cantidad == 0:
                            origen.delete()
                        else:
                            origen.save()
                    else:
                        return JsonResponse({'success': False, 'error': f'Stock insuficiente en {ubicacion_origen}'})
            
            # 2. Sumar al destino en PreInventario
            destino = PreInventario.objects.filter(
                codigo=codigo,
                producto=producto,
                categoria=categoria,
                proveedor=proveedor,
                ubicacion=ubicacion_destino
            ).first()
            
            if destino:
                destino.cantidad += cantidad
                destino.save()
            else:
                PreInventario.objects.create(
                    codigo=codigo,
                    producto=producto,
                    categoria=categoria,
                    proveedor=proveedor,
                    cantidad=cantidad,
                    ubicacion=ubicacion_destino
                )
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})
    



@login_required
def venta_detalle(request, id):
    venta = get_object_or_404(Venta, id=id)
    detalles = venta.detalles.all()
    abonos = Abono.objects.filter(venta=venta).order_by('-fecha')
    
    from decimal import Decimal
    from django.db.models import Sum
    
    abonos_total = Abono.objects.filter(venta=venta).aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
    pagado = float(abonos_total) if venta.tipo_pago == 'credito' else float(venta.total)
    saldo = float(venta.total) - pagado
    esta_pagado = saldo <= 0.01
    
    data = {
        'id': venta.id,
        'numero_factura': venta.numero_factura,
        'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
        'cliente': venta.cliente.nombre if venta.cliente else 'Consumidor Final',
        'nit': venta.cliente.nit if venta.cliente else '',
        'direccion': venta.cliente.direccion if venta.cliente else '',
        'tipo_documento': venta.tipo_documento,
        'tipo_pago': venta.tipo_pago,
        'subtotal': float(venta.subtotal),
        'total': float(venta.total),
        'pagado': round(pagado, 2),
        'saldo': round(saldo, 2),
        'esta_pagado': esta_pagado,
        'detalles': [{
            'producto': d.producto,
            'codigo': d.codigo,
            'cantidad': d.cantidad,
            'precio_unitario': float(d.precio_unitario),
            'subtotal': float(d.subtotal)
        } for d in detalles],
        'abonos': [{
            'fecha': a.fecha.strftime('%d/%m/%Y'),
            'cantidad': float(a.cantidad),
            'comentario': a.comentario or ''
        } for a in abonos]
    }
    return JsonResponse(data)


@login_required
def venta_abonos(request, id):
    venta = get_object_or_404(Venta, id=id)
    abonos = Abono.objects.filter(venta=venta).order_by('-fecha')
    
    data = [{
        'id': a.id,
        'fecha': a.fecha.strftime('%d/%m/%Y %H:%M'),
        'cantidad': float(a.cantidad),
        'comentario': a.comentario or ''
    } for a in abonos]
    return JsonResponse(data, safe=False)




@login_required
def registrar_abono(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            venta = get_object_or_404(Venta, id=data['venta_id'])
            
            from decimal import Decimal
            from django.db.models import Sum
            from datetime import datetime
            
            abonos_total = Abono.objects.filter(venta=venta).aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
            saldo = Decimal(str(venta.total)) - abonos_total
            
            cantidad_abono = Decimal(str(data['cantidad']))
            
            # Permitir hasta el saldo exacto (con tolerancia de 0.01)
            if cantidad_abono > saldo + Decimal('0.01'):
                return JsonResponse({'success': False, 'error': f'El abono no puede superar el saldo pendiente (Q{float(saldo):.2f})'})
            
            if cantidad_abono <= 0:
                return JsonResponse({'success': False, 'error': 'La cantidad debe ser mayor a cero'})
            
            # Si la cantidad es mayor al saldo, ajustar al saldo
            if cantidad_abono > saldo:
                cantidad_abono = saldo
            
            fecha_abono = datetime.strptime(data['fecha_abono'], '%Y-%m-%d').date()
            
            abono = Abono.objects.create(
                venta=venta,
                fecha=fecha_abono,
                cantidad=cantidad_abono,
                comentario=data.get('comentario', ''),
                usuario=request.user
            )
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def exportar_ventas(request, formato):
    ventas_list = Venta.objects.all().order_by('-fecha')
    
    busqueda = request.GET.get('busqueda', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if busqueda:
        ventas_list = ventas_list.filter(
            Q(numero_factura__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(cliente__nit__icontains=busqueda)
        )
    
    if fecha_desde:
        ventas_list = ventas_list.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        ventas_list = ventas_list.filter(fecha__date__lte=fecha_hasta)
    
    if formato == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        headers = ['Factura', 'Fecha', 'Cliente', 'NIT', 'Dirección', 'Tipo Doc', 'Tipo Pago', 'Subtotal', 'IVA', 'Total', 'Usuario']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, venta in enumerate(ventas_list, 2):
            ws.cell(row=row, column=1, value=venta.numero_factura)
            ws.cell(row=row, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=3, value=venta.cliente.nombre if venta.cliente else 'Consumidor Final')
            ws.cell(row=row, column=4, value=venta.cliente.nit if venta.cliente else '')
            ws.cell(row=row, column=5, value=venta.cliente.direccion if venta.cliente else '')
            ws.cell(row=row, column=6, value='Factura' if venta.tipo_documento == 'factura' else 'Envío')
            ws.cell(row=row, column=7, value=venta.tipo_pago.upper())
            ws.cell(row=row, column=8, value=float(venta.subtotal))
            #ws.cell(row=row, column=9, value=float(venta.iva))
            ws.cell(row=row, column=10, value=float(venta.total))
            ws.cell(row=row, column=11, value=venta.usuario.username)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'
        wb.save(response)
        return response
    
    elif formato == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="ventas.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Ventas", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(" ", styles['Normal']))
        
        data = [['Factura', 'Fecha', 'Cliente', 'Tipo Pago', 'Total', 'Usuario']]
        
        for venta in ventas_list:
            data.append([
                venta.numero_factura,
                venta.fecha.strftime('%d/%m/%Y'),
                venta.cliente.nombre[:30] if venta.cliente else 'Consumidor Final',
                venta.tipo_pago.upper(),
                f"Q{venta.total}",
                venta.usuario.username
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
    
    return JsonResponse({'error': 'Formato no soportado'}, status=400)


@login_required
def venta_detalle_pdf(request, id):
    venta = get_object_or_404(Venta, id=id)
    detalles = venta.detalles.all()
    abonos = Abono.objects.filter(venta=venta).order_by('-fecha')
    
    from decimal import Decimal
    from django.db.models import Sum
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    abonos_total = Abono.objects.filter(venta=venta).aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
    pagado = float(abonos_total) if venta.tipo_pago == 'credito' else float(venta.total)
    saldo = float(venta.total) - pagado
    esta_pagado = saldo <= 0
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="venta_{venta.numero_factura}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, alignment=1)
    
    # Título
    titulo = "FACTURA" if venta.tipo_documento == 'factura' else "ENVÍO"
    title = Paragraph(f"{titulo} #{venta.numero_factura}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Datos del cliente
    cliente_text = f"""
    <b>Cliente:</b> {venta.cliente.nombre if venta.cliente else 'Consumidor Final'}<br/>
    <b>NIT:</b> {venta.cliente.nit if venta.cliente else 'CF'}<br/>
    <b>Dirección:</b> {venta.cliente.direccion if venta.cliente else '-'}<br/>
    <b>Fecha:</b> {venta.fecha.strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Tipo Documento:</b> {titulo}<br/>
    <b>Tipo Pago:</b> {venta.tipo_pago.upper()}<br/>
    <b>Estado:</b> {'Pagado' if esta_pagado else 'Pendiente'}
    """
    cliente_para = Paragraph(cliente_text, styles['Normal'])
    elements.append(cliente_para)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Tabla de productos
    data = [['Cantidad', 'Producto', 'Precio', 'Subtotal']]
    for d in detalles:
        data.append([str(d.cantidad), d.producto, f"Q{d.precio_unitario}", f"Q{d.subtotal}"])
    
    table = Table(data, colWidths=[0.8*inch, 3.5*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Totales (sin IVA)
    totales_text = f"""
    <b>TOTAL:</b> Q{venta.total}
    """
    if venta.tipo_pago == 'credito':
        totales_text += f"""<br/>
        <b>Pagado:</b> Q{pagado:.2f}<br/>
        <b>Saldo Pendiente:</b> Q{saldo:.2f}
        """
    totales_para = Paragraph(totales_text, styles['Normal'])
    elements.append(totales_para)
    
    # Tabla de abonos si es crédito
    if venta.tipo_pago == 'credito' and abonos:
        elements.append(Spacer(1, 0.2 * inch))
        abono_title = Paragraph("Historial de Abonos", styles['Heading4'])
        elements.append(abono_title)
        
        abono_data = [['Fecha', 'Cantidad', 'Comentario']]
        for a in abonos:
            abono_data.append([a.fecha.strftime('%d/%m/%Y'), f"Q{a.cantidad}", a.comentario or '-'])
        
        abono_table = Table(abono_data, colWidths=[1.5*inch, 1.2*inch, 2.5*inch])
        abono_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(abono_table)
    
    doc.build(elements)
    return response
    venta = get_object_or_404(Venta, id=id)
    detalles = venta.detalles.all()
    abonos = Abono.objects.filter(venta=venta).order_by('-fecha')
    
    from decimal import Decimal
    from django.db.models import Sum
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    abonos_total = Abono.objects.filter(venta=venta).aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
    pagado = float(abonos_total) if venta.tipo_pago == 'credito' else float(venta.total)
    saldo = float(venta.total) - pagado
    esta_pagado = saldo <= 0
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="venta_{venta.numero_factura}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, alignment=1)
    
    # Título
    titulo = "FACTURA" if venta.tipo_documento == 'factura' else "ENVÍO"
    title = Paragraph(f"{titulo} #{venta.numero_factura}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Datos del cliente
    cliente_text = f"""
    <b>Cliente:</b> {venta.cliente.nombre if venta.cliente else 'Consumidor Final'}<br/>
    <b>NIT:</b> {venta.cliente.nit if venta.cliente else 'CF'}<br/>
    <b>Dirección:</b> {venta.cliente.direccion if venta.cliente else '-'}<br/>
    <b>Fecha:</b> {venta.fecha.strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Tipo Documento:</b> {titulo}<br/>
    <b>Tipo Pago:</b> {venta.tipo_pago.upper()}<br/>
    <b>Estado:</b> {'Pagado' if esta_pagado else 'Pendiente'}
    """
    cliente_para = Paragraph(cliente_text, styles['Normal'])
    elements.append(cliente_para)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Tabla de productos
    data = [['Cantidad', 'Producto', 'Precio', 'Subtotal']]
    for d in detalles:
        data.append([str(d.cantidad), d.producto, f"Q{d.precio_unitario}", f"Q{d.subtotal}"])
    
    table = Table(data, colWidths=[0.8*inch, 3.5*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Totales
    totales_text = f"""
    <b>Subtotal:</b> Q{venta.subtotal}<br/>
    <b>IVA (12%):</b> Q{venta.iva}<br/>
    <b>TOTAL:</b> Q{venta.total}
    """
    if venta.tipo_pago == 'credito':
        totales_text += f"""<br/>
        <b>Pagado:</b> Q{pagado:.2f}<br/>
        <b>Saldo Pendiente:</b> Q{saldo:.2f}
        """
    totales_para = Paragraph(totales_text, styles['Normal'])
    elements.append(totales_para)
    
    # Tabla de abonos si es crédito
    if venta.tipo_pago == 'credito' and abonos:
        elements.append(Spacer(1, 0.2 * inch))
        abono_title = Paragraph("Historial de Abonos", styles['Heading4'])
        elements.append(abono_title)
        
        abono_data = [['Fecha', 'Cantidad', 'Comentario']]
        for a in abonos:
            abono_data.append([a.fecha.strftime('%d/%m/%Y'), f"Q{a.cantidad}", a.comentario or '-'])
        
        abono_table = Table(abono_data, colWidths=[1.5*inch, 1.2*inch, 2.5*inch])
        abono_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(abono_table)
    
    doc.build(elements)
    return response


    venta = get_object_or_404(Venta, id=id)
    detalles = venta.detalles.all()
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factura_{venta.numero_factura}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, alignment=1)
    
    # Título
    title = Paragraph(f"FACTURA / ENVÍO #{venta.numero_factura}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Datos del cliente
    cliente_text = f"<b>Cliente:</b> {venta.cliente.nombre if venta.cliente else 'Consumidor Final'}<br/>"
    cliente_text += f"<b>NIT:</b> {venta.cliente.nit if venta.cliente else 'CF'}<br/>"
    cliente_text += f"<b>Dirección:</b> {venta.cliente.direccion if venta.cliente else '-'}<br/>"
    cliente_text += f"<b>Fecha:</b> {venta.fecha.strftime('%d/%m/%Y %H:%M')}<br/>"
    cliente_text += f"<b>Tipo Pago:</b> {venta.tipo_pago.upper()}"
    
    cliente_para = Paragraph(cliente_text, styles['Normal'])
    elements.append(cliente_para)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Tabla de productos
    data = [['Cantidad', 'Producto', 'Precio', 'Subtotal']]
    for d in detalles:
        data.append([str(d.cantidad), d.producto, f"Q{d.precio_unitario}", f"Q{d.subtotal}"])
    
    table = Table(data, colWidths=[0.8*inch, 3.5*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Totales
    totales = f"""
    <b>Subtotal:</b> Q{venta.subtotal}<br/>
    <b>IVA (12%):</b> Q{venta.iva}<br/>
    <b>TOTAL:</b> Q{venta.total}
    """
    totales_para = Paragraph(totales, styles['Normal'])
    elements.append(totales_para)
    
    doc.build(elements)
    return response




@login_required
def inventario_general(request):
    # Agrupar por código, producto, categoría, proveedor y sumar cantidades
    from django.db.models import Sum
    
    inventario = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).annotate(
        total_stock=Sum('cantidad')
    )
    
    # Ordenamiento
    orden = request.GET.get('orden', 'codigo')  # por defecto ordenar por código
    if orden == 'codigo':
        inventario = inventario.order_by('codigo')
    elif orden == 'producto':
        inventario = inventario.order_by('producto')
    elif orden == '-codigo':
        inventario = inventario.order_by('-codigo')
    elif orden == '-producto':
        inventario = inventario.order_by('-producto')
    else:
        inventario = inventario.order_by('codigo')
    
    # Filtros de búsqueda
    busqueda = request.GET.get('busqueda', '')
    
    if busqueda:
        inventario = inventario.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    # Calcular estado para cada producto
    resultados = []
    for item in inventario:
        stock = item['total_stock']
        
        if stock <= 2:
            estado_color = 'danger'
            estado_texto = 'Crítico'
        elif stock <= 6:
            estado_color = 'warning'
            estado_texto = 'Bajo'
        else:
            estado_color = 'success'
            estado_texto = 'Normal'
        
        resultados.append({
            'codigo': item['codigo'],
            'producto': item['producto'],
            'categoria': item['categoria'],
            'proveedor': item['proveedor'] or '-',
            'cantidad': stock,
            'estado_color': estado_color,
            'estado_texto': estado_texto
        })
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [20, 50, 100]:
            por_pagina = 20
    except:
        por_pagina = 20
    
    paginator = Paginator(resultados, por_pagina)
    page_number = request.GET.get('page')
    productos_paginados = paginator.get_page(page_number)
    
    context = {
        'productos': productos_paginados,
        'busqueda': busqueda,
        'orden': orden,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/inventario_general.html', context)


@login_required
def exportar_inventario_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    
    # Agrupar datos
    inventario = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).annotate(
        total_stock=Sum('cantidad')
    ).order_by('codigo', 'producto')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        inventario = inventario.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_general.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Inventario General - Stock Consolidado", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(" ", styles['Normal']))
    
    data = [['Código', 'Producto', 'Categoría', 'Proveedor', 'Stock Total', 'Estado']]
    
    for item in inventario:
        stock = item['total_stock']
        if stock <= 2:
            estado = "Crítico"
        elif stock <= 6:
            estado = "Bajo"
        else:
            estado = "Normal"
        
        data.append([
            item['codigo'], 
            item['producto'], 
            item['categoria'], 
            item['proveedor'] or '-', 
            str(stock), 
            estado
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_inventario_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Agrupar datos
    inventario = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).annotate(
        total_stock=Sum('cantidad')
    ).order_by('codigo', 'producto')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        inventario = inventario.filter(
            Q(codigo__icontains=busqueda) |
            Q(producto__icontains=busqueda) |
            Q(categoria__icontains=busqueda) |
            Q(proveedor__icontains=busqueda)
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario General"
    
    headers = ['Código', 'Producto', 'Categoría', 'Proveedor', 'Stock Total', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, item in enumerate(inventario, 2):
        stock = item['total_stock']
        if stock <= 2:
            estado = "Crítico"
        elif stock <= 6:
            estado = "Bajo"
        else:
            estado = "Normal"
        
        ws.cell(row=row, column=1, value=item['codigo'])
        ws.cell(row=row, column=2, value=item['producto'])
        ws.cell(row=row, column=3, value=item['categoria'])
        ws.cell(row=row, column=4, value=item['proveedor'] or '-')
        ws.cell(row=row, column=5, value=stock)
        ws.cell(row=row, column=6, value=estado)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_general.xlsx"'
    wb.save(response)
    return response


@login_required
def listar_clientes(request):
    clientes = Cliente.objects.all().order_by('nombre')
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        clientes = clientes.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(direccion__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(telefono2__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(observaciones__icontains=busqueda)
        )
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [20, 50, 100]:
            por_pagina = 20
    except:
        por_pagina = 20
    
    paginator = Paginator(clientes, por_pagina)
    page_number = request.GET.get('page')
    clientes_paginados = paginator.get_page(page_number)
    
    context = {
        'clientes': clientes_paginados,
        'busqueda': busqueda,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/clientes.html', context)


@login_required
def crear_cliente(request):
    if request.method == 'POST':
        try:
            cliente = Cliente.objects.create(
                nit=request.POST.get('nit'),
                nombre=request.POST.get('nombre'),
                direccion=request.POST.get('direccion'),
                telefono=request.POST.get('telefono'),
                telefono2=request.POST.get('telefono2'),
                email=request.POST.get('email'),
                observaciones=request.POST.get('observaciones')
            )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    
    if request.method == 'GET':
        data = {
            'id': cliente.id,
            'nit': cliente.nit,
            'nombre': cliente.nombre,
            'direccion': cliente.direccion or '',
            'telefono': cliente.telefono or '',
            'telefono2': cliente.telefono2 or '',
            'email': cliente.email or '',
            'observaciones': cliente.observaciones or '',
        }
        return JsonResponse(data)
    
    if request.method == 'POST':
        try:
            cliente.nit = request.POST.get('nit')
            cliente.nombre = request.POST.get('nombre')
            cliente.direccion = request.POST.get('direccion')
            cliente.telefono = request.POST.get('telefono')
            cliente.telefono2 = request.POST.get('telefono2')
            cliente.email = request.POST.get('email')
            cliente.observaciones = request.POST.get('observaciones')
            cliente.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def eliminar_cliente(request, id):
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, id=id)
        cliente.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def exportar_clientes_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    
    clientes = Cliente.objects.all().order_by('nombre')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        clientes = clientes.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(direccion__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Directorio de Clientes", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(" ", styles['Normal']))
    
    data = [['NIT', 'Nombre', 'Dirección', 'Teléfono', 'Email', 'Observaciones']]
    
    for c in clientes:
        data.append([
            c.nit, 
            c.nombre, 
            (c.direccion or '')[:40], 
            c.telefono or '', 
            c.email or '',
            (c.observaciones or '')[:30]
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_clientes_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    clientes = Cliente.objects.all().order_by('nombre')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        clientes = clientes.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(direccion__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    headers = ['NIT', 'Nombre', 'Dirección', 'Teléfono', 'Teléfono 2', 'Email', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, c in enumerate(clientes, 2):
        ws.cell(row=row, column=1, value=c.nit)
        ws.cell(row=row, column=2, value=c.nombre)
        ws.cell(row=row, column=3, value=c.direccion or '')
        ws.cell(row=row, column=4, value=c.telefono or '')
        ws.cell(row=row, column=5, value=c.telefono2 or '')
        ws.cell(row=row, column=6, value=c.email or '')
        ws.cell(row=row, column=7, value=c.observaciones or '')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response


@login_required
def listar_proveedores(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        proveedores = proveedores.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(persona_contacto__icontains=busqueda) |
            Q(ciudad__icontains=busqueda) |
            Q(pais__icontains=busqueda)
        )
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [20, 50, 100]:
            por_pagina = 20
    except:
        por_pagina = 20
    
    paginator = Paginator(proveedores, por_pagina)
    page_number = request.GET.get('page')
    proveedores_paginados = paginator.get_page(page_number)
    
    context = {
        'proveedores': proveedores_paginados,
        'busqueda': busqueda,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/proveedores.html', context)


@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        try:
            proveedor = Proveedor.objects.create(
                nit=request.POST.get('nit'),
                nombre=request.POST.get('nombre'),
                telefono=request.POST.get('telefono'),
                email=request.POST.get('email'),
                persona_contacto=request.POST.get('persona_contacto'),
                direccion_fiscal=request.POST.get('direccion_fiscal'),
                ciudad=request.POST.get('ciudad'),
                pais=request.POST.get('pais'),
                codigo_postal=request.POST.get('codigo_postal'),
                observaciones=request.POST.get('observaciones')
            )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    
    if request.method == 'GET':
        data = {
            'id': proveedor.id,
            'nit': proveedor.nit,
            'nombre': proveedor.nombre,
            'telefono': proveedor.telefono or '',
            'email': proveedor.email or '',
            'persona_contacto': proveedor.persona_contacto or '',
            'direccion_fiscal': proveedor.direccion_fiscal or '',
            'ciudad': proveedor.ciudad or '',
            'pais': proveedor.pais or '',
            'codigo_postal': proveedor.codigo_postal or '',
            'observaciones': proveedor.observaciones or '',
        }
        return JsonResponse(data)
    
    if request.method == 'POST':
        try:
            proveedor.nit = request.POST.get('nit')
            proveedor.nombre = request.POST.get('nombre')
            proveedor.telefono = request.POST.get('telefono')
            proveedor.email = request.POST.get('email')
            proveedor.persona_contacto = request.POST.get('persona_contacto')
            proveedor.direccion_fiscal = request.POST.get('direccion_fiscal')
            proveedor.ciudad = request.POST.get('ciudad')
            proveedor.pais = request.POST.get('pais')
            proveedor.codigo_postal = request.POST.get('codigo_postal')
            proveedor.observaciones = request.POST.get('observaciones')
            proveedor.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def eliminar_proveedor(request, id):
    if request.method == 'POST':
        proveedor = get_object_or_404(Proveedor, id=id)
        proveedor.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def exportar_proveedores_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        proveedores = proveedores.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="proveedores.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Directorio de Proveedores", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(" ", styles['Normal']))
    
    data = [['NIT', 'Nombre', 'Teléfono', 'Email', 'Contacto', 'Ciudad', 'País']]
    
    for p in proveedores:
        data.append([
            p.nit, 
            p.nombre, 
            p.telefono or '', 
            p.email or '',
            p.persona_contacto or '',
            p.ciudad or '',
            p.pais or ''
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_proveedores_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    proveedores = Proveedor.objects.all().order_by('nombre')
    
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        proveedores = proveedores.filter(
            Q(nit__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    headers = ['NIT', 'Nombre', 'Teléfono', 'Email', 'Persona Contacto', 'Dirección Fiscal', 'Ciudad', 'País', 'Código Postal']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, p in enumerate(proveedores, 2):
        ws.cell(row=row, column=1, value=p.nit)
        ws.cell(row=row, column=2, value=p.nombre)
        ws.cell(row=row, column=3, value=p.telefono or '')
        ws.cell(row=row, column=4, value=p.email or '')
        ws.cell(row=row, column=5, value=p.persona_contacto or '')
        ws.cell(row=row, column=6, value=p.direccion_fiscal or '')
        ws.cell(row=row, column=7, value=p.ciudad or '')
        ws.cell(row=row, column=8, value=p.pais or '')
        ws.cell(row=row, column=9, value=p.codigo_postal or '')
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="proveedores.xlsx"'
    wb.save(response)
    return response



import json
from django.core.serializers.json import DjangoJSONEncoder


@login_required
def catalogo(request):
    # Agrupar por código, producto, categoría, proveedor (evitar duplicados)
    from django.db.models import Max
    
    productos_raw = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).annotate(
        max_id=Max('id')
    ).distinct()
    
    categorias_dict = {}
    for p in productos_raw:
        # Obtener un registro de ejemplo para descripción e imagen
        ejemplo = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            categoria=p['categoria'],
            proveedor=p['proveedor']
        ).first()
        
        cat = p['categoria'] if p['categoria'] else 'Otros'
        if cat not in categorias_dict:
            categorias_dict[cat] = []
        
        # Sumar stock de todas las ubicaciones de este producto
        stock_total = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            proveedor=p['proveedor']
        ).aggregate(total=models.Sum('cantidad'))['total'] or 0
        
        categorias_dict[cat].append({
            'codigo': p['codigo'],
            'nombre': p['producto'],
            'descripcion': ejemplo.descripcion if ejemplo else '',
            'imagen_catalogo': ejemplo.imagen_catalogo.url if ejemplo and ejemplo.imagen_catalogo else '',
            'precio': '',  # Precio opcional
            'stock': stock_total
        })
    
    context = {
        'categorias_json': json.dumps(categorias_dict, cls=DjangoJSONEncoder),
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/catalogo.html', context)



@login_required
def exportar_catalogo_pdf(request):
    from weasyprint import HTML
    from django.template.loader import get_template
    from django.db.models import Sum
    
    # Agrupar por código, producto, categoría, proveedor (evitar duplicados)
    productos_raw = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).distinct()
    
    categorias_dict = {}
    for p in productos_raw:
        # Obtener un registro de ejemplo para imagen y descripción
        ejemplo = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            categoria=p['categoria'],
            proveedor=p['proveedor']
        ).first()
        
        cat = p['categoria'] if p['categoria'] else 'Otros'
        if cat not in categorias_dict:
            categorias_dict[cat] = []
        
        # Sumar stock de todas las ubicaciones (opcional)
        stock_total = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            proveedor=p['proveedor']
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        categorias_dict[cat].append({
            'codigo': p['codigo'],
            'producto': p['producto'],
            'descripcion': ejemplo.descripcion if ejemplo else '',
            'imagen_catalogo': ejemplo.imagen_catalogo if ejemplo and ejemplo.imagen_catalogo else None,
            'precio': '',
            'stock': stock_total
        })
    
    # Renderizar template HTML
    template = get_template('core/catalogo_pdf.html')
    html_string = template.render({
        'categorias': categorias_dict,
        'nombre_usuario': request.user.username,
        'request': request,
    })
    
    # Generar PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="catalogo_el_charro.pdf"'
    
    HTML(string=html_string).write_pdf(response)
    return response

"""
@login_required
def catalogo_pedidos(request):
    # Obtener combinaciones únicas de producto + proveedor (sin descripcion)
    productos_raw = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).distinct()
    
    resultados = []
    for p in productos_raw:
        # Obtener la descripción de cualquier registro (la primera)
        any_registro = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            proveedor=p['proveedor']
        ).first()
        
        descripcion = any_registro.descripcion if any_registro else ''
        
        # Obtener ubicaciones y stock
        ubicaciones = list(PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            proveedor=p['proveedor']
        ).values('ubicacion', 'cantidad'))
        
        stock_total = sum(ub['cantidad'] for ub in ubicaciones)
        
        # Buscar imagen
        imagen_registro = PreInventario.objects.filter(
            codigo=p['codigo'],
            producto=p['producto'],
            proveedor=p['proveedor']
        ).exclude(imagen_catalogo__isnull=True).exclude(imagen_catalogo='').first()
        
        imagen_url = None
        if imagen_registro and imagen_registro.imagen_catalogo:
            try:
                imagen_url = imagen_registro.imagen_catalogo.url
            except:
                imagen_url = None
        
        resultados.append({
            'codigo': p['codigo'],
            'producto': p['producto'],
            'categoria': p['categoria'],
            'proveedor': p['proveedor'],
            'descripcion': descripcion,
            'imagen_catalogo': imagen_url,
            'ubicaciones': ubicaciones,
            'stock_total': stock_total
        })
    
    pendientes_count = Pedido.objects.filter(estado='pendiente').count()
    
    context = {
        'productos': resultados,
        'pendientes_count': pendientes_count,
        'nombre_usuario': request.user.username,
    }
    
    return render(request, 'core/catalogo_pedidos.html', context)

"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import PreInventario, Pedido


@login_required
@grupo_requerido('Administrador', 'Operativo')
def catalogo_pedidos(request):
    # 1. Obtener combinaciones únicas de productos
    productos_raw = PreInventario.objects.values(
        'codigo', 'producto', 'categoria', 'proveedor'
    ).distinct()
    
    resultados = []
    categorias_set = set() # Para los botones de categorías
    
    for p in productos_raw:
        if p['categoria']:
            categorias_set.add(p['categoria'])

        # Filtro para detalles y stock total
        base_qs = PreInventario.objects.filter(
            codigo=p['codigo'], producto=p['producto'], proveedor=p['proveedor']
        )

        any_reg = base_qs.first()
        descripcion = any_reg.descripcion if any_reg else ''
        
        # Ubicaciones para el Modal
        ubicaciones = list(base_qs.values('ubicacion', 'cantidad'))
        stock_total = sum(u['cantidad'] for u in ubicaciones)
        
        # Imagen del catálogo
        img_reg = base_qs.exclude(imagen_catalogo__isnull=True).exclude(imagen_catalogo='').first()
        img_url = img_reg.imagen_catalogo.url if img_reg and img_reg.imagen_catalogo else None
        
        resultados.append({
            'codigo': p['codigo'],
            'producto': p['producto'],
            'categoria': p['categoria'],
            'proveedor': p['proveedor'],
            'descripcion': descripcion,
            'imagen_catalogo': img_url,
            'ubicaciones': ubicaciones,
            'stock_total': stock_total
        })
    
    context = {
        'productos': resultados,
        'categorias': sorted(list(categorias_set)),
        'pendientes_count': Pedido.objects.filter(estado='pendiente').count(),
    }
    return render(request, 'core/catalogo_pedidos.html', context)



@login_required
def obtener_ubicaciones_producto(request):
    codigo = request.GET.get('codigo')
    producto = request.GET.get('producto')
    
    ubicaciones = PreInventario.objects.filter(
        codigo=codigo,
        producto=producto,
        cantidad__gt=0
    ).values('ubicacion', 'cantidad')
    
    # Precio sugerido (puedes ajustarlo)
    precio = 100  # Temporal
    
    data = [{
        'ubicacion': u['ubicacion'],
        'stock': u['cantidad'],
        'precio': float(precio)
    } for u in ubicaciones]
    
    return JsonResponse(data, safe=False)



@login_required
def checkout_pedido(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            carrito = data.get('carrito', [])
            cliente_data = data.get('cliente', {})
            
            if not carrito:
                return JsonResponse({'success': False, 'error': 'Carrito vacío'})
            
            # Buscar o crear cliente
            cliente = None
            cliente_id = cliente_data.get('id')
            
            if cliente_id and cliente_id != 'null' and cliente_id != '':
                try:
                    cliente = Cliente.objects.filter(id=int(cliente_id)).first()
                except:
                    cliente = None
            
            if not cliente:
                nit = cliente_data.get('nit', '')
                nombre = cliente_data.get('nombre', '')
                
                # Si es Consumidor Final o CF
                if nit == 'CF' or nombre == 'Consumidor Final':
                    cliente = Cliente.objects.filter(nit='CF').first()
                    if not cliente:
                        cliente = Cliente.objects.create(
                            nit='CF',
                            nombre='Consumidor Final'
                        )
                elif nit and nit != '':
                    # Buscar por NIT
                    cliente = Cliente.objects.filter(nit=nit).first()
                    if not cliente and nombre:
                        cliente = Cliente.objects.create(
                            nit=nit,
                            nombre=nombre,
                            direccion=cliente_data.get('direccion', '')
                        )
                elif nombre:
                    # Cliente sin NIT
                    cliente = Cliente.objects.create(
                        nit=None,
                        nombre=nombre,
                        direccion=cliente_data.get('direccion', '')
                    )
            
            # Si aún no hay cliente, crear uno por defecto
            if not cliente:
                cliente = Cliente.objects.create(
                    nit='CF',
                    nombre='Consumidor Final'
                )
            
            # Generar número de pedido
            ultimo_pedido = Pedido.objects.order_by('-id').first()
            if ultimo_pedido:
                num_pedido = str(int(ultimo_pedido.numero_pedido) + 1).zfill(8)
            else:
                num_pedido = '00000001'
            
            # Calcular totales
            subtotal = 0
            for item in carrito:
                subtotal += item['cantidad'] * item['precio']
            total = subtotal
            
            # Crear pedido
            pedido = Pedido.objects.create(
                numero_pedido=num_pedido,
                cliente=cliente,
                direccion_envio=cliente_data.get('direccion', ''),
                estado='pendiente',
                subtotal=subtotal,
                total=total,
                usuario=request.user
            )
            
            # Crear detalles del pedido
            for item in carrito:
                DetallePedido.objects.create(
                    pedido=pedido,
                    producto=item['producto'],
                    codigo=item['codigo'],
                    proveedor=item.get('proveedor', ''),
                    cantidad=item['cantidad'],
                    precio_unitario=item['precio'],
                    ubicacion=item['ubicacion'],
                    subtotal=item['cantidad'] * item['precio']
                )
            
            return JsonResponse({'success': True, 'pedido_id': pedido.id, 'numero_pedido': num_pedido})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    pendientes_count = Pedido.objects.filter(usuario=request.user, estado='pendiente').count()
    
    context = {
        'nombre_usuario': request.user.username,
        'pendientes_count': pendientes_count,
    }
    return render(request, 'core/checkout_pedido.html', context)



@login_required
def historial_pedidos(request):
    pedidos_list = Pedido.objects.all().order_by('-fecha')
    
    # Filtros
    busqueda = request.GET.get('busqueda', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if busqueda:
        pedidos_list = pedidos_list.filter(
            Q(numero_pedido__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(cliente__nit__icontains=busqueda)
        )
    
    if fecha_desde:
        pedidos_list = pedidos_list.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos_list = pedidos_list.filter(fecha__date__lte=fecha_hasta)
    
    # Paginación
    por_pagina = request.GET.get('por_pagina', 20)
    try:
        por_pagina = int(por_pagina)
        if por_pagina not in [20, 50, 100]:
            por_pagina = 20
    except:
        por_pagina = 20
    
    paginator = Paginator(pedidos_list, por_pagina)
    page_number = request.GET.get('page')
    pedidos = paginator.get_page(page_number)
    
    # Contar pedidos pendientes (sin paginación)
    pendientes_count = Pedido.objects.filter(estado='pendiente').count()
    
    context = {
        'pedidos': pedidos,
        'pendientes_count': pendientes_count,
        'busqueda': busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'por_pagina': por_pagina,
        'nombre_usuario': request.user.username,
    }
    return render(request, 'core/historial_pedidos.html', context)


@login_required
def detalle_pedido(request, id):
    pedido = get_object_or_404(Pedido, id=id)
    detalles = pedido.detalles.all()
    
    data = {
        'id': pedido.id,
        'numero_pedido': pedido.numero_pedido,
        'fecha': pedido.fecha.strftime('%d/%m/%Y %H:%M'),
        'cliente': pedido.cliente.nombre if pedido.cliente else 'Sin cliente',
        'direccion': pedido.direccion_envio or '-',
        'usuario': pedido.usuario.username,
        'estado': pedido.estado,
        'tipo_documento': pedido.tipo_documento or '',
        'tipo_pago': pedido.tipo_pago or '',
        'numero_factura': pedido.numero_factura or '',
        'observaciones': pedido.observaciones or '',
        'subtotal': float(pedido.subtotal),
        'total': float(pedido.total),
        'detalles': [{
            'id': d.id,
            'producto': d.producto,
            'codigo': d.codigo,
            'cantidad': d.cantidad,
            'precio_unitario': float(d.precio_unitario),
            'subtotal': float(d.subtotal),
            'ubicacion': d.ubicacion
        } for d in detalles]
    }
    return JsonResponse(data)



@login_required
def actualizar_campos_pedido(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pedido = get_object_or_404(Pedido, id=data['pedido_id'])
            
            pedido.tipo_documento = data.get('tipo_documento')
            pedido.tipo_pago = data.get('tipo_pago')
            pedido.numero_factura = data.get('numero_factura')
            pedido.observaciones = data.get('observaciones')
            pedido.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})





@login_required
def finalizar_pedido(request, id):
    if request.method == 'POST':
        try:
            pedido = get_object_or_404(Pedido, id=id)
            
            if not pedido.tipo_documento or not pedido.tipo_pago or not pedido.numero_factura:
                return JsonResponse({'success': False, 'error': 'Complete todos los campos obligatorios'})
            
            if Venta.objects.filter(numero_factura=pedido.numero_factura).exists():
                return JsonResponse({'success': False, 'error': f'El número de factura {pedido.numero_factura} ya existe. Use otro número.'})
            
            # DESCONTAR STOCK del PreInventario por cada detalle (incluyendo proveedor)
            for detalle in pedido.detalles.all():
                pre = PreInventario.objects.filter(
                    codigo=detalle.codigo,
                    producto=detalle.producto,
                    proveedor=detalle.proveedor,  # Incluir proveedor
                    ubicacion=detalle.ubicacion
                ).first()
                
                if pre:
                    if pre.cantidad >= detalle.cantidad:
                        pre.cantidad -= detalle.cantidad
                        if pre.cantidad == 0:
                            pre.delete()
                        else:
                            pre.save()
                    else:
                        return JsonResponse({'success': False, 'error': f'Stock insuficiente para {detalle.producto} ({detalle.proveedor}) en {detalle.ubicacion}'})
                else:
                    return JsonResponse({'success': False, 'error': f'Producto {detalle.producto} ({detalle.proveedor}) no encontrado en {detalle.ubicacion}'})
            
            # Crear venta en el módulo Ventas
            venta = Venta.objects.create(
                numero_factura=pedido.numero_factura,
                cliente=pedido.cliente,
                subtotal=pedido.subtotal,
                total=pedido.total,
                tipo_documento=pedido.tipo_documento,
                tipo_pago=pedido.tipo_pago,
                comentario=f"Pedido #{pedido.numero_pedido} - {pedido.observaciones or ''}",
                usuario=request.user
            )
            
            # Crear detalles de venta
            for detalle in pedido.detalles.all():
                DetalleVenta.objects.create(
                    venta=venta,
                    producto=detalle.producto,
                    codigo=detalle.codigo,
                    proveedor=detalle.proveedor or '',
                    cantidad=detalle.cantidad,
                    precio_unitario=detalle.precio_unitario,
                    subtotal=detalle.subtotal
                )
            
            # Cambiar estado del pedido a completado
            pedido.estado = 'completado'
            pedido.save()
            
            return JsonResponse({'success': True, 'venta_id': venta.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})




@login_required
def actualizar_precio_detalle(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            detalle = get_object_or_404(DetallePedido, id=data['detalle_id'])
            
            detalle.precio_unitario = data['precio']
            detalle.subtotal = detalle.cantidad * detalle.precio_unitario
            detalle.save()
            
            # Actualizar total del pedido
            pedido = detalle.pedido
            nuevo_total = sum(d.subtotal for d in pedido.detalles.all())
            pedido.subtotal = nuevo_total
            pedido.total = nuevo_total
            pedido.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def eliminar_detalle_pedido(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            detalle = get_object_or_404(DetallePedido, id=data['detalle_id'])
            pedido = detalle.pedido
            
            # Solo permitir eliminar si el pedido está pendiente
            if pedido.estado != 'pendiente':
                return JsonResponse({'success': False, 'error': 'No se puede eliminar un pedido ya completado'})
            
            # Verificar que el usuario sea el dueño del pedido
            if pedido.usuario != request.user:
                return JsonResponse({'success': False, 'error': 'No autorizado'})
            
            detalle.delete()
            
            # Recalcular totales del pedido
            nuevo_subtotal = sum(d.subtotal for d in pedido.detalles.all())
            pedido.subtotal = nuevo_subtotal
            pedido.total = nuevo_subtotal
            pedido.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@login_required
def anular_pedido(request, id):
    if request.method == 'POST':
        try:
            pedido = get_object_or_404(Pedido, id=id)
            
            if pedido.estado != 'pendiente':
                return JsonResponse({'success': False, 'error': 'Solo se pueden anular pedidos pendientes'})
            
            pedido.estado = 'anulado'
            pedido.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def crear_categoria(request):
    if request.method == 'POST':
        try:
            from core.models import Categoria
            nombre = request.POST.get('nombre').strip().lower()
            
            # Verificar si ya existe
            categoria_existente = Categoria.objects.filter(nombre=nombre).first()
            if categoria_existente:
                return JsonResponse({'success': True, 'nombre': categoria_existente.nombre})
            
            # Crear nueva categoría
            categoria = Categoria.objects.create(
                nombre=nombre,
                descripcion=request.POST.get('descripcion', '')
            )
            return JsonResponse({'success': True, 'nombre': categoria.nombre})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def test_debug(request):
    from django.conf import settings
    import json
    return JsonResponse({
        'debug': settings.DEBUG,
        'allowed_hosts': settings.ALLOWED_HOSTS,
    })